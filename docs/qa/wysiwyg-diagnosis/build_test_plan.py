"""
Generatore del test plan WYSIWYG diagnosis come .xlsx.

Output: test-plan-template.xlsx — pronto per upload su Google Drive
        (Google Sheets importa .xlsx nativamente).

Eseguire da repo root: python3 docs/qa/wysiwyg-diagnosis/build_test_plan.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT_PATH = "docs/qa/wysiwyg-diagnosis/test-plan-template.xlsx"

COLOR_HEADER_BG = "1B2B4B"
COLOR_HEADER_FG = "FFFFFF"
COLOR_SUBHEADER_BG = "E5E0D5"
COLOR_SUBHEADER_FG = "2D2D2D"
COLOR_INPUT_BG = "FAFAF8"
COLOR_INSTRUCTIONS_BG = "F2F0EA"

THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)


def style_header(cell, bold=True):
    cell.font = Font(name="Arial", size=11, bold=bold, color=COLOR_HEADER_FG)
    cell.fill = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER


def style_input_cell(cell):
    cell.font = Font(name="Arial", size=10)
    cell.fill = PatternFill(start_color=COLOR_INPUT_BG, end_color=COLOR_INPUT_BG, fill_type="solid")
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    cell.border = THIN_BORDER


def style_readonly_cell(cell, bold=False):
    cell.font = Font(name="Arial", size=10, bold=bold, color=COLOR_SUBHEADER_FG)
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    cell.border = THIN_BORDER


def style_instructions(cell, bold=False, size=10):
    cell.font = Font(name="Arial", size=size, bold=bold, color="2D2D2D")
    cell.fill = PatternFill(start_color=COLOR_INSTRUCTIONS_BG, end_color=COLOR_INSTRUCTIONS_BG, fill_type="solid")
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)


def build_sheet_inizia(wb):
    ws = wb.active
    ws.title = "Inizia da qui"
    ws.column_dimensions["A"].width = 110

    ws["A1"] = "DIAGNOSI WYSIWYG — Test Plan per Elena"
    ws["A1"].font = Font(name="Arial", size=18, bold=True, color=COLOR_HEADER_BG)
    ws.row_dimensions[1].height = 32

    ws["A2"] = "Studio Legale Saltelli WordPress · Adsolut SRLS · 2026-05-05"
    ws["A2"].font = Font(name="Arial", size=10, italic=True, color="6B6B6B")
    ws.row_dimensions[2].height = 18

    rows = [
        ("", 8),
        ("PERCHÉ STAI FACENDO QUESTO TEST", 22),
        (
            "Il sito Saltelli su staging è stato costruito con un nuovo modello CMS che dovrebbe permetterti di "
            "modificare i contenuti in autonomia, senza chiamare il team tecnico. "
            "Vogliamo capire dove questo modello funziona e dove invece ti fa perdere tempo o ti confonde. "
            "Il tuo feedback servirà a sistemare le cose PRIMA che il sito vada in produzione.",
            68,
        ),
        ("", 8),
        ("REGOLE DEL TEST (importanti, leggi tutto)", 22),
        (
            "1. NON studiare il manuale prima. Apri WP-Admin e prova. Il manuale puoi consultarlo SOLO se ti blocchi "
            "davvero (e in quel caso annota nella colonna 'Cosa hai fatto alla fine' che hai dovuto consultarlo).",
            54,
        ),
        (
            "2. Tieni il timer. Per ogni scenario, segna i minuti che hai impiegato — anche quelli persi a cercare "
            "dove andare. È il dato più importante del test.",
            42,
        ),
        (
            "3. Se non riesci a fare uno scenario, NON è un fallimento tuo. È un fallimento del sistema che abbiamo "
            "costruito. Segna 'No' nella colonna 'Riuscita?' e annota cosa stavi cercando.",
            48,
        ),
        (
            "4. Sii onesta sulla frustrazione. Se uno scenario ti ha fatto innervosire, segna 5. Non è un giudizio "
            "personale — è un dato che ci serve per capire dove intervenire.",
            42,
        ),
        (
            "5. Aggiungi note dove vuoi. La colonna 'Note libere' è per qualsiasi pensiero — anche cose come "
            "'mi aspettavo che il tasto fosse rosso e invece è grigio'. Tutto è utile.",
            42,
        ),
        ("", 8),
        ("PRIMA DI INIZIARE, FAI QUESTI 4 STEP", 22),
        (
            "1. Apri il foglio 'Test Plan' (in basso, scheda 2). Lì sono i 10 scenari da fare in ordine.",
            22,
        ),
        (
            "2. Apri in un'altra finestra del browser: https://staging.studiolegalesaltelli.it/wp-admin/ "
            "(login con le credenziali che ti ha mandato Duccio).",
            38,
        ),
        (
            "3. Apri in una terza finestra il sito pubblico: https://staging.studiolegalesaltelli.it/ "
            "(per verificare che le tue modifiche si vedano).",
            38,
        ),
        (
            "4. Tieni a portata di mano un cronometro (anche quello del telefono va bene).",
            22,
        ),
        ("", 8),
        ("TEMPO TOTALE STIMATO", 22),
        (
            "Circa 50-90 minuti se il sistema funziona bene. Se finisci in meno di 50 minuti, sei una macchina. "
            "Se ne impieghi più di 2 ore, fermati a metà e annota dove sei rimasta — aggregheremo lo stesso.",
            42,
        ),
        ("", 8),
        ("DOPO IL TEST", 22),
        (
            "Compila anche il foglio 'Open feedback' (scheda 3) — sono 6 domande aperte sul tuo rapporto generale "
            "con WordPress, oltre agli scenari specifici.",
            42,
        ),
        (
            "Quando hai finito, fammi sapere via Slack o email a tech@adsolut.it. "
            "Aggregeremo i risultati e capiremo come sistemare le cose.",
            38,
        ),
        ("", 8),
        ("DOMANDE? CONTATTO", 22),
        (
            "Per dubbi sul test stesso (NON sul WP-Admin): scrivi a Duccio Santoro su Slack o email tech@adsolut.it. "
            "Per problemi tecnici tipo 'non riesco a entrare in WP-Admin': stesso canale.",
            42,
        ),
        ("", 8),
        ("Grazie. Il tuo lavoro qui vale tantissimo per chiudere bene il progetto.", 22),
    ]

    row = 3
    for content, height in rows:
        cell = ws.cell(row=row, column=1, value=content)
        if content and content == content.upper() and len(content) > 5 and not content.startswith("Studio"):
            cell.font = Font(name="Arial", size=12, bold=True, color=COLOR_HEADER_BG)
        else:
            style_instructions(cell, bold=False)
        ws.row_dimensions[row].height = height
        row += 1

    return ws


def build_sheet_test_plan(wb):
    ws = wb.create_sheet("Test Plan")

    widths = {
        "A": 5,
        "B": 50,
        "C": 14,
        "D": 14,
        "E": 14,
        "F": 14,
        "G": 38,
        "H": 38,
        "I": 38,
        "J": 30,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    headers = [
        "#",
        "Cosa devi fare",
        "Tempo target (min)",
        "Tempo impiegato (min)",
        "Frustrazione (1-5)",
        "Riuscita?",
        "Cosa cercavi e non trovavi",
        "Cosa ti aspettavi",
        "Cosa hai fatto alla fine",
        "Note libere",
    ]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        style_header(cell)
    ws.row_dimensions[1].height = 34
    ws.freeze_panes = "B2"

    scenarios = [
        (
            "1",
            "Lo Studio cambia operatore telefonico. Cambia il numero di telefono pubblico mostrato sul sito da "
            "+39 081 1813 1119 a +39 081 9999 9999. Verifica che il numero sia aggiornato sul footer del sito "
            "pubblico (apri la home in un'altra finestra). Quando hai finito, RIMETTILO al valore originale "
            "+39 081 1813 1119 — è un test, non vogliamo lasciare il numero finto.",
            2,
        ),
        (
            "2",
            "Aggiungi una nuova FAQ sul tema 'cartelle esattoriali'. La domanda è: 'Posso ricorrere contro una "
            "cartella che ho già pagato?'. La risposta è una breve frase tipo: 'Sì, il pagamento non preclude il "
            "ricorso. Hai 60 giorni dalla notifica per impugnare.' Pubblicala e verifica che appaia su /faq/ del "
            "sito pubblico.",
            5,
        ),
        (
            "3",
            "Il cliente vuole modificare il testo della pagina /lo-studio/. In particolare, vuole cambiare il "
            "brand statement (la frase di apertura tipo 'Un atelier legale italiano. Quattro avvocati a Chiaia...'). "
            "Trova dove modificarlo e cambia 'Vent'anni di pratica' in 'Oltre vent'anni di pratica'. Verifica sulla "
            "pagina pubblica /lo-studio/ che il testo sia aggiornato.",
            5,
        ),
        (
            "4",
            "Aggiorna la bio estesa di Avv. Emiliano Saltelli. Trova la sua scheda nel WP-Admin, individua il campo "
            "della bio estesa e aggiungi questa frase alla fine della bio attuale: 'Membro del Comitato Tecnico "
            "dell'Associazione Avvocati Tributaristi Campani.' Verifica sulla pagina pubblica "
            "/avvocati/emiliano-saltelli/ che la nuova frase sia visibile.",
            5,
        ),
        (
            "5",
            "Sulla pagina /costi/ del sito pubblico vedi 3 box 'Modalità di consulenza' (in presenza, video, parere). "
            "Modifica il titolo del primo box (quello 'in presenza') aggiungendo '— Studio Chiaia' al titolo "
            "esistente. Verifica che la modifica si veda su /costi/.",
            3,
        ),
        (
            "6",
            "Aggiungi un nuovo caso vinto rappresentativo del 2024. Tribunale di Napoli, civile, recupero crediti "
            "per fornitura non saldata, importo 75.000 €. Inseriscilo anonimizzato come saresti tu a giudicare "
            "appropriato (niente nomi cliente, niente importi specifici). Verifica che appaia su /casi/.",
            8,
        ),
        (
            "7",
            "Modifica l'hero della pagina /faq/. L'eyebrow attuale è '§ Risorse · Domande frequenti' — cambialo "
            "in '§ Risorse · FAQ legali'. Lascia tutto il resto invariato. Verifica sulla pagina pubblica.",
            3,
        ),
        (
            "8",
            "Lo Studio decide di aggiungere una 20a area di pratica: 'Diritto agroalimentare'. Crea la pagina "
            "dell'area come faresti normalmente — slug, titolo, descrizione breve, e quello che ti sembra il minimo "
            "indispensabile per pubblicarla in modo sensato. Lascia il toggle Tier-1 SPENTO (è un'area Tier-2). "
            "Verifica che appaia su /competenze/ e che sia accessibile cliccandoci sopra.",
            10,
        ),
        (
            "9",
            "Cambia il payoff del logo dello Studio. Quello attuale è 'Diritto, con misura.' — cambialo "
            "temporaneamente in 'Diritto, su misura.' (solo per test). Verifica che il payoff aggiornato si veda "
            "sull'header del sito pubblico (sotto al logo). Quando hai finito, rimettilo a 'Diritto, con misura.'",
            2,
        ),
        (
            "10",
            "Il cliente ti manda una guida PDF da pubblicare nelle Guide gratuite. Per il test usa un PDF qualsiasi "
            "che hai sul tuo computer (anche un PDF stupido, basta che sia .pdf). Crea una nuova guida intitolata "
            "'Test guida cartelle 2026', categoria 'tributario', carica il PDF, e verifica che appaia su "
            "/guide-gratuite/ del sito pubblico.",
            8,
        ),
    ]

    riuscita_dv = DataValidation(type="list", formula1='"Sì,Sì con difficoltà,No,Non provata"', allow_blank=True)
    riuscita_dv.error = "Scegli una delle opzioni"
    riuscita_dv.errorTitle = "Valore non valido"
    ws.add_data_validation(riuscita_dv)

    frustr_dv = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True)
    frustr_dv.error = "Inserisci un valore tra 1 e 5"
    frustr_dv.errorTitle = "Valore non valido"
    ws.add_data_validation(frustr_dv)

    for idx, (n, descr, target_min) in enumerate(scenarios, start=2):
        ws.cell(row=idx, column=1, value=n)
        ws.cell(row=idx, column=2, value=descr)
        ws.cell(row=idx, column=3, value=target_min)
        ws.cell(row=idx, column=4, value="")
        ws.cell(row=idx, column=5, value="")
        ws.cell(row=idx, column=6, value="")
        ws.cell(row=idx, column=7, value="")
        ws.cell(row=idx, column=8, value="")
        ws.cell(row=idx, column=9, value="")
        ws.cell(row=idx, column=10, value="")

        for col in [1, 2, 3]:
            style_readonly_cell(ws.cell(row=idx, column=col))
        ws.cell(row=idx, column=1).font = Font(name="Arial", size=11, bold=True, color=COLOR_HEADER_BG)
        ws.cell(row=idx, column=1).alignment = Alignment(horizontal="center", vertical="top")
        ws.cell(row=idx, column=3).alignment = Alignment(horizontal="center", vertical="top")

        for col in range(4, 11):
            style_input_cell(ws.cell(row=idx, column=col))

        riuscita_dv.add(f"F{idx}")
        frustr_dv.add(f"E{idx}")

        ws.row_dimensions[idx].height = 100

    total_row = len(scenarios) + 3
    ws.cell(row=total_row, column=2, value="TOTALE TEMPO IMPIEGATO")
    ws.cell(row=total_row, column=2).font = Font(name="Arial", size=11, bold=True)
    ws.cell(row=total_row, column=3, value=f"=SUM(C2:C{len(scenarios)+1})")
    ws.cell(row=total_row, column=3).font = Font(name="Arial", size=11, bold=True)
    ws.cell(row=total_row, column=4, value=f"=SUM(D2:D{len(scenarios)+1})")
    ws.cell(row=total_row, column=4).font = Font(name="Arial", size=11, bold=True)
    ws.cell(row=total_row, column=4).fill = PatternFill(start_color="FFE599", end_color="FFE599", fill_type="solid")

    return ws


def build_sheet_open_feedback(wb):
    ws = wb.create_sheet("Open feedback")

    ws.column_dimensions["A"].width = 80
    ws.column_dimensions["B"].width = 80

    ws["A1"] = "Domanda"
    ws["B1"] = "La tua risposta"
    style_header(ws["A1"])
    style_header(ws["B1"])
    ws.row_dimensions[1].height = 30

    questions = [
        (
            "1. Durante il test, ti sono successe cose che ti hanno fatto perdere tempo ma che non rientravano "
            "in nessuno scenario specifico? (es. menù che non trovavi, messaggi WP poco chiari, lentezza, ecc.)"
        ),
        (
            "2. Pensando al sito Saltelli IN GENERALE (anche prima di questo test), quali sono le 3 cose che ti "
            "fanno perdere più tempo o ti fanno innervosire di più quando devi modificarlo?"
        ),
        (
            "3. Wishlist: se potessi cambiare 3 cose del WP-Admin del sito Saltelli per renderlo più facile "
            "da usare per te, quali sarebbero?"
        ),
        (
            "4. Pensando a un sito web 'ideale' su cui modificare contenuti facilmente, c'è qualcosa che ti "
            "aspetteresti di trovare in modo ovvio e che invece su Saltelli non trovi?"
        ),
        (
            "5. C'è un caso ricorrente in cui finisci sempre per chiamare il team tecnico anche se penseresti "
            "di poter fare da sola? Quale?"
        ),
        (
            "6. Su una scala da 1 (per niente) a 5 (totalmente), quanto ti senti AUTONOMA nel modificare i "
            "contenuti del sito Saltelli oggi? E perché?"
        ),
    ]

    for idx, q in enumerate(questions, start=2):
        ws.cell(row=idx, column=1, value=q)
        ws.cell(row=idx, column=2, value="")
        style_readonly_cell(ws.cell(row=idx, column=1))
        style_input_cell(ws.cell(row=idx, column=2))
        ws.row_dimensions[idx].height = 130

    return ws


def build_sheet_riferimenti(wb):
    ws = wb.create_sheet("Riferimenti rapidi")

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 90

    ws["A1"] = "Riferimento"
    ws["B1"] = "Valore"
    style_header(ws["A1"])
    style_header(ws["B1"])
    ws.row_dimensions[1].height = 30

    rows = [
        ("URL WP-Admin staging", "https://staging.studiolegalesaltelli.it/wp-admin/"),
        ("URL sito pubblico staging", "https://staging.studiolegalesaltelli.it/"),
        ("Credenziali", "Username e password te li ha mandati Duccio separatamente (NON inserirle in questo foglio)"),
        ("Manuale editoriale completo", "https://github.com/Adsolut-Ai-Agency/saltelli-wp/blob/main/docs/EDITOR-HANDOFF.md"),
        ("Architettura tecnica (per curiosità)", "https://github.com/Adsolut-Ai-Agency/saltelli-wp/blob/main/docs/ARCHITECTURE.md"),
        ("Contatto per dubbi sul test", "Duccio Santoro · Slack o tech@adsolut.it"),
        ("", ""),
        ("Importante", "Il sito staging NON è il sito live del cliente. Il dominio reale studiolegalesaltelli.it punta ancora al vecchio sito Elementor. Tutto quello che fai qui è SOLO test — niente di quello che modifichi qui finisce sul sito visibile al cliente o al pubblico."),
    ]

    for idx, (label, value) in enumerate(rows, start=2):
        ws.cell(row=idx, column=1, value=label)
        ws.cell(row=idx, column=2, value=value)
        if label == "Importante":
            ws.cell(row=idx, column=1).font = Font(name="Arial", size=10, bold=True, color="C0392B")
            ws.cell(row=idx, column=2).font = Font(name="Arial", size=10, color="2D2D2D")
            ws.cell(row=idx, column=2).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.row_dimensions[idx].height = 60
        else:
            style_readonly_cell(ws.cell(row=idx, column=1), bold=True)
            style_readonly_cell(ws.cell(row=idx, column=2))
            ws.row_dimensions[idx].height = 22

    return ws


def main():
    wb = Workbook()
    build_sheet_inizia(wb)
    build_sheet_test_plan(wb)
    build_sheet_open_feedback(wb)
    build_sheet_riferimenti(wb)
    wb.save(OUTPUT_PATH)
    print(f"OK · scritto {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
